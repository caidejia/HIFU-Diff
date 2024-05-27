import json
import math
import os
from sklearn.utils import resample
import numpy as np
import torch
from PIL import Image
import matplotlib.pyplot as plt
from ssim import SSIM
from tqdm import tqdm
from scipy import stats
import openpyxl
target_path_sig = ['E:/Data/DL/0706/Image/Hilbert/label', 'E:/Data/DL/0717/Image/Hilbert/label',
                 'E:/Data/DL/0828/Image/Hilbert/label', ':/Data/DL/0829/Image/Hilbert/label',
                 'E:/Data/DL/0905/Image/Hilbert/label', 'E:/Data/DL/1009/Image/Hilbert/label','E:/Data/DL/unseen/Image/Hilbert/label']

target_path_21A = ['E:/Data/DL/0706/Image_21A/Hilbert/label', 'E:/Data/DL/0717/Image_21A/Hilbert/label',
                 'E:/Data/DL/0828/Image_21A/Hilbert/label', 'E:/Data/DL/0829/Image_21A/Hilbert/label',
                 'E:/Data/DL/0905/Image_21A/Hilbert/label', 'E:/Data/DL/1009/Image_21A/Hilbert/label','E:/Data/DL/unseen/Image_21A/Hilbert/label']


input_path = r'E:\Data\dataset_pre\DDIM\UNSEEN\Image\eta=0.5_ts=5_ir=-57'.replace("\\", "/")
compare_path = r'E:\Data\dataset_pre\Image\UNSEEN\FUSNET1_21A_60_2'.replace("\\", "/")
isUnseen = True
doCmp = True


def calculate_psnr(img1, img2, border=0):
    # img1 and img2 have range [0, 255]
    if not img1.shape == img2.shape:
        raise ValueError('Input images must have the same dimensions.')
    h, w = img1.shape[:2]
    img1 = img1[border:h - border, border:w - border]
    img2 = img2[border:h - border, border:w - border]
    img1 = img1.astype(np.float64)
    img2 = img2.astype(np.float64)
    mse = np.mean((img1 - img2) ** 2)
    if mse == 0:
        return float('inf')
    return 20 * math.log10(255.0 / math.sqrt(mse))


def calculate_ssim(img1, img2):
    img1 = torch.from_numpy(img1)
    img2 = torch.from_numpy(img2)
    img1 = img1 / torch.max(torch.abs(img1))
    img2 = img2 / torch.max(torch.abs(img2))
    img1 = img1[None, None, :, :].permute(0, 1, 3, 2)
    img2 = img2[None, None, :, :].permute(0, 1, 3, 2)
    ssim_f = SSIM()
    aa = ssim_f(img2, img1)
    return aa


def check_border(img):
    t = 0
    b = img.shape[0]
    r = img.shape[1]
    l = 0
    # img shape is (656,875)
    for j in range(img.shape[0]):  # top
        if not np.all(img[j, :] == 255):
            t = j
            break
    for j in range(img.shape[0] - 1, 0, -1):  # bottom
        if not np.all(img[j, :] == 255):
            b = j
            break
    for j in range(img.shape[1]):
        if not np.all(img[:, j] == 255):
            l = j
            break
    for j in range(img.shape[1] - 1, 0, -1):
        if not np.all(img[:, j] == 255):
            r = j
            break
    # print(f'height={b-t},width={r-l}')
    return t, b, l, r

def caculate_p_value(data1, data2):
    # 首先，测试数据是否符合正态分布
    # 如果两个p-value都大于0.05，那么我们可以假设数据是正态分布的。
    # print(stats.normaltest(data1))
    # print(stats.normaltest(data2))

    # 使用独立样本t检验计算p-value
    # 使用这个检验之前，最好确认数据满足正态分布和方差齐性
    t_stat, p_value = stats.ttest_ind(data1, data2)
    print(f'T-test p-value: {p_value}')
    #
    # # 如果数据不符合正态分布，可以使用Mann-Whitney U检验
    # u_stat, p_value = stats.mannwhitneyu(data1, data2, alternative='two-sided')
    # print(f'Mann-Whitney U-test p-value: {p_value}')


def Caculate_CIs(value):

    # 自助法置信区间计算
    bootstrap_samples = 1000
    alpha = 0.05
    bootstrapped_value_means = []

    for _ in range(bootstrap_samples):
        boot_sample = resample(value, replace=True)
        boot_mean = np.mean(boot_sample)
        bootstrapped_value_means.append(boot_mean)

    # 计算百分位数
    lower_percentile = np.percentile(bootstrapped_value_means, 100 * (alpha/2))
    upper_percentile = np.percentile(bootstrapped_value_means, 100 * (1 - (alpha/2)))

    print(f"{(1-alpha)*100}% confidence interval for the data values: [{lower_percentile}, {upper_percentile}]")

def idx_image_dataset():
    index_21A = True
    files = os.listdir(input_path)
    len_f = len(files)
    step = 0
    row_index = step + 1
    sum_psnr = 0
    sum_ssim = 0
    sum_psnr_c = 0
    sum_ssim_c = 0
    ssim_record = []
    psnr_record = []
    ssim_record_2 = []
    psnr_record_2 = []
    file_path = '../store_pre/res.txt'
    if doCmp:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

    # prefixes = ['0717']
    # mat_files = [f for f in files if any(f.startswith(p) for p in prefixes)]
    # len_f = len(mat_files)
    with open(file_path, 'w') as file_t:
        with tqdm(total=len_f, unit='img') as pbar:
            for file in files:
                bg = file[0:4]
                flag=True
                if not index_21A:
                    if isUnseen:
                        target_path = target_path_sig[6]
                        flag = False
                    elif bg == '0706':
                        target_path = target_path_sig[0]
                    elif bg == '0828':
                        target_path = target_path_sig[2]
                    elif bg == '0829':
                        target_path = target_path_sig[3]
                    elif bg == '0905':
                        target_path = target_path_sig[4]
                    elif bg == '1009':
                        target_path = target_path_sig[5]
                    else:
                        target_path = target_path_sig[1]
                        if bg!='0717':
                            flag= False
                else:
                    if isUnseen:
                        target_path = target_path_21A[6]
                        flag =False
                    elif bg == '0706':
                        target_path = target_path_21A[0]
                    elif bg == '0828':
                        target_path = target_path_21A[2]
                    elif bg == '0829':
                        target_path = target_path_21A[3]
                    elif bg == '0905':
                        target_path = target_path_21A[4]
                    elif bg == '1009':
                        target_path = target_path_21A[5]
                    else:
                        target_path = target_path_21A[1]
                        if bg != '0717':
                            flag = False
                if flag:
                    ori_image = os.path.join(target_path, file[5:])
                else:
                    ori_image = os.path.join(target_path, file)
                pre_image = os.path.join(input_path, file)
                label, pre = Image.open(ori_image).convert('L'), Image.open(pre_image).convert('L')
                label, pre = np.array(label, dtype=np.float32), np.array(pre, dtype=np.float32)
                t, b, l, r = check_border(label)
                f1 = calculate_psnr(label[t:b, l:r], pre[t:b, l:r])
                f2 = calculate_ssim(label[t:b, l:r], pre[t:b, l:r])
                ssim_record.append(f2.item())
                psnr_record.append(f1)
                sum_psnr += f1
                sum_ssim += f2
                if doCmp:
                    try:
                        cmp_image = os.path.join(compare_path, file)
                        cmp = Image.open(cmp_image).convert('L')
                    except FileNotFoundError:
                        cmp_image = os.path.join(compare_path,'0717_'+file)
                        cmp = Image.open(cmp_image).convert('L')
                    cmp= np.array(cmp, dtype=np.float32)
                    t1 = calculate_psnr(label[t:b, l:r], cmp[t:b, l:r])
                    t2 = calculate_ssim(label[t:b, l:r], cmp[t:b, l:r])
                    ssim_record_2.append(t2.item())
                    psnr_record_2.append(t1)
                    sum_psnr_c += t1
                    sum_ssim_c += t2
                    strings = [file,str(f1), str(t1),'', str(f2.item()),str(t2.item())]
                    # 将四个字符串分别写入四个列中
                    for col_index, string in enumerate(strings, start=1):
                        # 在当前行的不同列中写入字符串
                        if row_index == 1:
                            ppp=[input_path,"",compare_path]
                            for idx, pp in enumerate(ppp, start=1):
                                sheet.cell(row=row_index, column=idx).value = pp
                            row_index += 1
                            titles = ['name', 'psnr_now', 'psnr_cmp', ' ', 'ssim_now', 'ssim_cmp']
                            for idx, title in enumerate(titles, start=1):
                                sheet.cell(row=row_index, column=idx).value = title
                            row_index += 1
                        sheet.cell(row=row_index, column=col_index).value = string
                row_index+=1
                pbar.update()
                step += 1
                temp_res = f'{file}      ssim_value:{f2}      psnr_value:{f1}\n'
                file_t.write(temp_res)
        msg1=f'avg psnr value:{sum_psnr / len_f}\n'
        msg2=f'avg ssim value:{sum_ssim / len_f}\n'
        file_t.write(msg1)
        file_t.write(msg2)
        print(f'avg psnr value:{sum_psnr / len_f}')
        print(f'avg ssim value:{sum_ssim / len_f}')
        if doCmp:
            avgs=['avg',str(sum_psnr / len_f),str(sum_psnr_c/len_f)," ",str((sum_ssim/len_f).item()),str((sum_ssim_c/len_f).item())]
            for col_index, avg in enumerate(avgs, start=1):
                sheet.cell(row=row_index, column=col_index).value = avg
            workbook.save('../store_pre/res.xlsx')
            print(f'cmp avg psnr value:{sum_psnr_c / len_f}')
            print(f'cmp avg ssim value:{sum_ssim_c / len_f}')

    fig, (ax1, ax2) = plt.subplots(2, 1)
    # 在第一个子图上绘制数据
    ax1.plot(ssim_record)
    ax1.set_xlabel('index')
    ax1.set_ylabel('ssim value')
    # 在第二个子图上绘制数据
    ax2.plot(psnr_record)
    ax2.set_xlabel('index')
    ax2.set_ylabel('psnr value')
    # 调整子图之间的间距
    plt.tight_layout()
    # 保存图形
    plt.savefig('../store_pre/res.png')
    # 显示图形
    plt.show()

    if len(ssim_record_2)!=0:
        print("p-value of ssim")
        caculate_p_value(ssim_record,ssim_record_2)
        print("p-value of psnr")
        caculate_p_value(psnr_record, psnr_record_2)
        print("CIs of ssim 1")
        Caculate_CIs(ssim_record)
        print("CIs of ssim 2")
        Caculate_CIs(ssim_record_2)
        print("CIs of psnr 1")
        Caculate_CIs(psnr_record)
        print("CIs of psnr 2")
        Caculate_CIs(psnr_record_2)




def idx_sigImage():
    root_dir = 'C:/Document/pycharm/predictor/store_ImageSig/psnr/'
    filename = 'pre.png'
    filename2 = 'label.png'
    dir_label = root_dir + filename2
    dir_pre = root_dir + filename
    label, pre = Image.open(dir_label).convert('L'), Image.open(dir_pre).convert('L')
    label, pre = np.array(label, dtype=np.float32), np.array(pre, dtype=np.float32)
    t, b, l, r = check_border(label)
    f1 = calculate_psnr(label[t:b, l:r], pre[t:b, l:r])
    f2 = calculate_ssim(label[t:b, l:r], pre[t:b, l:r])

    print(f'ROI psnr value:{f1}')
    print(f'ROI ssim value:{f2}')




if __name__ == '__main__':
    idx_image_dataset()
